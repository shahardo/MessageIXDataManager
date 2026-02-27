"""
Data Export Manager - Handles saving modified MESSAGEix data back to Excel files

Produces Excel files compatible with InputManager / ParameterParsingStrategy:
  - One worksheet per parameter, named after the parameter
  - Row 1: column headers  (dim1, dim2, ..., value)
  - Row 2+: data rows
  - Optional "Sets" sheet in combined-sets format
"""

import pandas as pd
import openpyxl
from openpyxl.styles import Font
import os
from typing import Callable, Dict, List, Optional
from core.data_models import ScenarioData, Parameter


class DataExportManager:
    """Manager for exporting MESSAGEix scenario data to Excel files."""

    def __init__(self):
        self.export_formats = {
            'xlsx': self._export_to_xlsx,
            'xls':  self._export_to_xlsx,   # xls → xlsx under the hood
        }

    # ------------------------------------------------------------------
    # Public API
    # ------------------------------------------------------------------

    def save_scenario(self, scenario: ScenarioData, file_path: str,
                      modified_only: bool = True,
                      progress_callback: Optional[Callable[[int, str], None]] = None) -> bool:
        """Save scenario data to an Excel file.

        The produced file is compatible with InputManager: each parameter gets
        its own worksheet (sheet name = parameter name).  modified_only is
        intentionally ignored — we always write the *complete* scenario so the
        output file is a valid, re-importable input file.

        Args:
            scenario: ScenarioData object to save.
            file_path: Destination path (will be created/overwritten).
            modified_only: Kept for API compatibility; always ignored.
            progress_callback: Optional ``(percent: int, message: str)`` callback
                called as each worksheet is written so the caller can update a
                progress bar.

        Returns:
            True on success, False on error.
        """
        try:
            _, ext = os.path.splitext(file_path)
            ext = ext.lower().lstrip('.')
            export_fn = self.export_formats.get(ext, self._export_to_xlsx)
            return export_fn(scenario, file_path, progress_callback)
        except Exception as e:
            print(f"Error saving scenario: {e}")
            return False

    def has_modified_data(self, scenario: ScenarioData) -> bool:
        """Return True if the scenario has any modified parameters."""
        return bool(scenario.modified)

    def get_modified_parameters_count(self, scenario: ScenarioData) -> int:
        """Return the number of modified parameters."""
        return len(scenario.modified)

    def clear_modified_flags(self, scenario: ScenarioData):
        """Clear all modified flags after a successful save."""
        scenario.modified.clear()
        scenario.change_history.clear()

    # ------------------------------------------------------------------
    # Internal helpers
    # ------------------------------------------------------------------

    def _export_to_xlsx(self, scenario: ScenarioData, file_path: str,
                        progress_callback: Optional[Callable[[int, str], None]] = None) -> bool:
        """Write the scenario to a MESSAGEix-compatible .xlsx workbook.

        Mirrors the input-file format so the saved file can be reloaded:
          - One sheet per set, named after the set; row 1 = bold set-name
            header (skipped by parser), rows 2+ = one value per row.
            Matches SetParsingStrategy._parse_individual_set_sheet.
          - One sheet per input parameter, named after the parameter (≤31 chars);
            row 1 = bold column headers, rows 2+ = data rows.
            Matches ParameterParsingStrategy._parse_individual_parameter_sheet.
          - Result variables, equations, and postprocessed metrics are skipped.

        progress_callback, if provided, is called as (percent, message) after
        each worksheet is written so the caller can update a progress bar.
        """
        try:
            wb = openpyxl.Workbook()
            wb.remove(wb.active)   # discard the default empty sheet

            # Shared set to track used sheet names across sets and parameters
            # so we never create duplicate worksheet titles.
            used_sheet_names: set[str] = set()

            # Count total items to write for accurate progress reporting
            input_params = [
                (n, p) for n, p in scenario.parameters.items()
                if not p.metadata.get('result_type') and p.df is not None and not p.df.empty
            ]
            sets = [
                (n, s) for n, s in sorted((scenario.sets or {}).items())
                if (list(s) if hasattr(s, '__iter__') else s)
            ]
            total = len(sets) + len(input_params) or 1
            done = 0

            def _report(sheet_name: str):
                nonlocal done
                done += 1
                if progress_callback:
                    pct = int(done / total * 100)
                    progress_callback(pct, f"Writing {sheet_name}…")

            # --- Individual sheet per set ----------------------------------
            for set_name, series in sets:
                values = list(series) if hasattr(series, '__iter__') else [series]
                if not values:
                    continue
                set_sheet_name = self._unique_sheet_name(set_name, used_sheet_names)
                ws_set = wb.create_sheet(title=set_sheet_name)
                # Row 1: bold header (skipped by parser which uses min_row=2).
                # Use the original A1 label stored in Series.name (e.g. "level"
                # for the "level_renewable" sheet) to mirror the source file.
                a1_label = getattr(series, 'name', None) or set_name
                ws_set.cell(row=1, column=1, value=str(a1_label)).font = Font(bold=True)
                for row_idx, val in enumerate(values, start=2):
                    ws_set.cell(row=row_idx, column=1, value=str(val))
                _report(set_sheet_name)

            # --- One sheet per input parameter -----------------------------
            for param_name, param in input_params:
                sheet_name = self._unique_sheet_name(param_name, used_sheet_names)
                ws = wb.create_sheet(title=sheet_name)
                self._write_parameter_sheet(ws, param.df)
                _report(sheet_name)

            # Final step: write file to disk (~remaining work)
            if progress_callback:
                progress_callback(99, f"Writing {os.path.basename(file_path)}…")
            wb.save(file_path)
            if progress_callback:
                progress_callback(100, "Saved")
            return True

        except Exception as e:
            print(f"Error exporting to Excel: {e}")
            import traceback; traceback.print_exc()
            return False

    def _write_parameter_sheet(self, ws, df: pd.DataFrame) -> None:
        """Write a single parameter DataFrame to worksheet.

        Row 1: bold column headers.
        Rows 2+: data (pandas NaN converted to None so Excel gets blank cells).
        """
        headers = list(df.columns)

        # Header row
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=str(header))
            cell.font = Font(bold=True)

        # Data rows — convert NaN → None so openpyxl writes empty cells
        for row_idx, row in enumerate(df.itertuples(index=False, name=None), start=2):
            for col_idx, value in enumerate(row, start=1):
                # pandas may give numpy scalar types; cast to plain Python types
                if pd.isna(value) if not isinstance(value, str) else False:
                    value = None
                elif hasattr(value, 'item'):   # numpy scalar → Python scalar
                    value = value.item()
                ws.cell(row=row_idx, column=col_idx, value=value)

    def _unique_sheet_name(self, param_name: str, used: set) -> str:
        """Return a sanitised, unique sheet name (≤31 chars)."""
        base = self._sanitize_sheet_name(param_name)
        name = base
        counter = 2
        while name in used:
            suffix = f"_{counter}"
            name = base[:31 - len(suffix)] + suffix
            counter += 1
        used.add(name)
        return name

    def _sanitize_sheet_name(self, name: str) -> str:
        """Sanitise a string to be a valid Excel sheet name (≤31 chars, no special chars)."""
        for ch in ('\\', '/', '?', '*', '[', ']', ':'):
            name = name.replace(ch, '_')
        return name[:31]
