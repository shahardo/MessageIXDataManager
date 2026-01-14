"""
Tests for Results Analyzer
"""

import pytest
import pandas as pd
import os
import sys
import tempfile
import numpy as np
from openpyxl import Workbook

# Add src directory to path for imports
sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..', 'src'))

from managers.results_analyzer import ResultsAnalyzer
from core.data_models import ScenarioData, Parameter


class TestResultsAnalyzer:
    """Test cases for ResultsAnalyzer"""

    @pytest.fixture
    def temp_results_file(self):
        """Create a temporary results Excel file"""
        wb = Workbook()

        # Create a variable results sheet
        ws_var = wb.active
        ws_var.title = "var_ACT"
        ws_var['A1'] = 'region'
        ws_var['B1'] = 'technology'
        ws_var['C1'] = 'year'
        ws_var['D1'] = 'value'

        ws_var['A2'] = 'region1'
        ws_var['B2'] = 'coal'
        ws_var['C2'] = 2020
        ws_var['D2'] = 100.5

        ws_var['A3'] = 'region1'
        ws_var['B3'] = 'solar'
        ws_var['C3'] = 2020
        ws_var['D3'] = 75.25

        # Create an equation results sheet
        ws_equ = wb.create_sheet("equ_BALANCE")
        ws_equ['A1'] = 'region'
        ws_equ['B1'] = 'year'
        ws_equ['C1'] = 'value'

        ws_equ['A2'] = 'region1'
        ws_equ['B2'] = 2020
        ws_equ['C2'] = 0.0

        ws_equ['A3'] = 'region1'
        ws_equ['B3'] = 2021
        ws_equ['C3'] = 0.0

        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
            wb.save(tmp.name)
            yield tmp.name
        os.unlink(tmp.name)

    def test_initialization(self):
        """Test ResultsAnalyzer initialization"""
        analyzer = ResultsAnalyzer()
        assert analyzer.results == []
        assert analyzer.loaded_file_paths == []
        assert analyzer.summary_stats == {}

    def test_load_results_file(self, temp_results_file):
        """Test loading a results file"""
        analyzer = ResultsAnalyzer()
        results = analyzer.load_results_file(temp_results_file)

        assert results is not None
        assert isinstance(results, ScenarioData)
        assert len(results.parameters) == 2  # var_ACT and equ_BALANCE
        assert 'var_ACT' in results.get_parameter_names()
        assert 'equ_BALANCE' in results.get_parameter_names()

    def test_parse_result_sheet_variable(self, temp_results_file):
        """Test parsing variable result sheets"""
        analyzer = ResultsAnalyzer()
        analyzer.load_results_file(temp_results_file)

        param = analyzer.get_result_data('var_ACT')
        assert param is not None
        assert len(param.df) == 2
        assert list(param.df.columns) == ['region', 'technology', 'year', 'value']
        assert param.df.iloc[0]['value'] == 100.5
        assert param.df.iloc[1]['technology'] == 'solar'

    def test_parse_result_sheet_equation(self, temp_results_file):
        """Test parsing equation result sheets"""
        analyzer = ResultsAnalyzer()
        analyzer.load_results_file(temp_results_file)

        param = analyzer.get_result_data('equ_BALANCE')
        assert param is not None
        assert len(param.df) == 2
        assert list(param.df.columns) == ['region', 'year', 'value']
        # Note: Zeros are preserved as zeros (not converted to NaN)
        assert param.df.iloc[0]['value'] == 0.0  # Should remain 0.0
        assert param.df.iloc[1]['year'] == 2021

    def test_get_summary_stats(self, temp_results_file):
        """Test getting summary statistics"""
        analyzer = ResultsAnalyzer()
        analyzer.load_results_file(temp_results_file)

        stats = analyzer.get_summary_stats()

        assert 'total_variables' in stats
        assert 'total_equations' in stats
        assert 'total_data_points' in stats
        assert 'result_sheets' in stats

        assert stats['total_variables'] == 1  # var_ACT
        assert stats['total_equations'] == 1  # equ_BALANCE
        assert stats['total_data_points'] == 4  # 2 + 2 rows
        assert len(stats['result_sheets']) == 2

    def test_get_result_data(self, temp_results_file):
        """Test getting specific result data"""
        analyzer = ResultsAnalyzer()
        analyzer.load_results_file(temp_results_file)

        param = analyzer.get_result_data('var_ACT')
        assert param is not None
        assert param.name == 'var_ACT'

        # Test nonexistent result
        nonexistent = analyzer.get_result_data('nonexistent')
        assert nonexistent is None

    def test_get_all_result_names(self, temp_results_file):
        """Test getting all result names"""
        analyzer = ResultsAnalyzer()
        analyzer.load_results_file(temp_results_file)

        names = analyzer.get_all_result_names()
        assert len(names) == 2
        assert 'var_ACT' in names
        assert 'equ_BALANCE' in names

    def test_prepare_chart_data_line(self, temp_results_file):
        """Test preparing line chart data"""
        analyzer = ResultsAnalyzer()
        analyzer.load_results_file(temp_results_file)

        chart_data = analyzer.prepare_chart_data('var_ACT', 'line')
        assert chart_data is not None
        assert chart_data['title'] == 'var_ACT Results'
        assert 'data' in chart_data
        assert len(chart_data['data']) > 0

    def test_prepare_chart_data_bar(self, temp_results_file):
        """Test preparing bar chart data"""
        analyzer = ResultsAnalyzer()
        analyzer.load_results_file(temp_results_file)

        chart_data = analyzer.prepare_chart_data('var_ACT', 'bar')
        assert chart_data is not None
        assert 'data' in chart_data

    def test_prepare_chart_data_nonexistent(self, temp_results_file):
        """Test preparing chart data for nonexistent result"""
        analyzer = ResultsAnalyzer()
        analyzer.load_results_file(temp_results_file)

        chart_data = analyzer.prepare_chart_data('nonexistent', 'line')
        assert chart_data is None

    def test_get_current_results_single_file(self, temp_results_file):
        """Test getting current results with single file"""
        analyzer = ResultsAnalyzer()
        analyzer.load_results_file(temp_results_file)

        current = analyzer.get_current_results()
        assert current is not None
        assert len(current.parameters) == 2

    def test_get_current_results_no_files(self):
        """Test getting current results when no files loaded"""
        analyzer = ResultsAnalyzer()

        current = analyzer.get_current_results()
        assert current is None

    def test_multiple_files(self):
        """Test loading multiple result files"""
        analyzer = ResultsAnalyzer()

        # Create first file
        wb1 = Workbook()
        ws1 = wb1.active
        ws1.title = "var_COST"
        ws1['A1'] = 'value'
        ws1['A2'] = 100

        # Create second file
        wb2 = Workbook()
        ws2 = wb2.active
        ws2.title = "var_REVENUE"
        ws2['A1'] = 'value'
        ws2['A2'] = 200

        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp1, \
             tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp2:

            wb1.save(tmp1.name)
            wb2.save(tmp2.name)

            analyzer.load_results_file(tmp1.name)
            analyzer.load_results_file(tmp2.name)

            # Check loaded files
            assert len(analyzer.loaded_file_paths) == 2
            assert len(analyzer.results) == 2

            # Check combined results
            combined = analyzer.get_current_results()
            assert combined is not None
            assert len(combined.parameters) == 2
            assert 'var_COST' in combined.get_parameter_names()
            assert 'var_REVENUE' in combined.get_parameter_names()

        os.unlink(tmp1.name)
        os.unlink(tmp2.name)

    def test_get_loaded_file_paths(self, temp_results_file):
        """Test getting loaded file paths"""
        analyzer = ResultsAnalyzer()
        analyzer.load_results_file(temp_results_file)

        paths = analyzer.get_loaded_file_paths()
        assert len(paths) == 1
        assert paths[0] == temp_results_file

    def test_get_results_by_file_path(self, temp_results_file):
        """Test getting results by file path"""
        analyzer = ResultsAnalyzer()
        analyzer.load_results_file(temp_results_file)

        results = analyzer.get_results_by_file_path(temp_results_file)
        assert results is not None
        assert len(results.parameters) == 2

        # Test nonexistent path
        nonexistent = analyzer.get_results_by_file_path('/nonexistent/path.xlsx')
        assert nonexistent is None

    def test_remove_file(self, temp_results_file):
        """Test removing a loaded file"""
        analyzer = ResultsAnalyzer()
        analyzer.load_results_file(temp_results_file)

        assert len(analyzer.loaded_file_paths) == 1
        assert len(analyzer.results) == 1

        result = analyzer.remove_file(temp_results_file)
        assert result is True
        assert len(analyzer.loaded_file_paths) == 0
        assert len(analyzer.results) == 0

    def test_remove_nonexistent_file(self):
        """Test removing a nonexistent file"""
        analyzer = ResultsAnalyzer()

        result = analyzer.remove_file('/nonexistent/path.xlsx')
        assert result is False

    def test_clear_results(self, temp_results_file):
        """Test clearing all results"""
        analyzer = ResultsAnalyzer()
        analyzer.load_results_file(temp_results_file)

        assert len(analyzer.loaded_file_paths) == 1

        analyzer.clear_results()
        assert len(analyzer.loaded_file_paths) == 0
        assert len(analyzer.results) == 0
        assert analyzer.get_current_results() is None

    def test_is_result_sheet(self):
        """Test identifying result sheets"""
        analyzer = ResultsAnalyzer()

        # Create test workbook
        wb = Workbook()
        ws1 = wb.active
        ws1.title = "var_TEST"
        ws1['A1'] = 'header1'
        ws1['B1'] = 'header2'
        ws1['A2'] = 'data1'
        ws1['B2'] = 123

        ws2 = wb.create_sheet("input_data")
        ws2['A1'] = 'parameter'
        ws2['B1'] = 'value'

        # Test var_ sheet
        assert analyzer._is_result_sheet(ws1) is True

        # Test input sheet
        assert analyzer._is_result_sheet(ws2) is False

    def test_parse_result_sheet_with_none_headers(self):
        """Test parsing sheets with None headers"""
        analyzer = ResultsAnalyzer()

        wb = Workbook()
        ws = wb.active
        ws['A1'] = None  # None header
        ws['B1'] = 'value'
        ws['A2'] = 2020  # Year-like data
        ws['B2'] = 100

        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
            wb.save(tmp.name)

            analyzer.load_results_file(tmp.name)

            # Should have parsed the sheet
            results = analyzer.get_current_results()
            assert results is not None
            assert len(results.parameters) >= 1

        os.unlink(tmp.name)

    def test_file_not_found(self):
        """Test handling of nonexistent files"""
        analyzer = ResultsAnalyzer()

        with pytest.raises(FileNotFoundError):
            analyzer.load_results_file('/nonexistent/file.xlsx')

    def test_invalid_excel_file(self):
        """Test handling of invalid Excel files"""
        analyzer = ResultsAnalyzer()

        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
            tmp.write(b"This is not Excel data")
            tmp_path = tmp.name

        try:
            with pytest.raises(ValueError):
                analyzer.load_results_file(tmp_path)
        finally:
            os.unlink(tmp_path)

    def test_metadata_creation(self, temp_results_file):
        """Test that metadata is correctly created for results"""
        analyzer = ResultsAnalyzer()
        analyzer.load_results_file(temp_results_file)

        var_param = analyzer.get_result_data('var_ACT')
        assert var_param.metadata['result_type'] == 'variable'

        equ_param = analyzer.get_result_data('equ_BALANCE')
        assert equ_param.metadata['result_type'] == 'equation'

        # Check other metadata
        assert var_param.metadata['units'] == 'N/A'
        assert 'ACT' in var_param.metadata['desc']
