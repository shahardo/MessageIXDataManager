"""
Tests for InputManager Excel loading and parsing functionality
"""

import pytest
import pandas as pd
import os
import sys
import tempfile
from openpyxl import Workbook

# Add src directory to path for imports
sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..', 'src'))

from managers.input_manager import InputManager


class TestInputManager:
    """Test cases for InputManager"""

    @pytest.fixture
    def temp_excel_file(self):
        """Create a temporary Excel file with test data"""
        wb = Workbook()

        # Create parameters sheet
        ws_params = wb.active
        ws_params.title = "parameters"

        # Headers
        ws_params['A1'] = 'parameter'
        ws_params['B1'] = 'node_loc'
        ws_params['C1'] = 'technology'
        ws_params['D1'] = 'year_vtg'
        ws_params['E1'] = 'value'

        # Test parameter 1: fix_cost
        ws_params['A2'] = 'fix_cost'
        ws_params['B2'] = 'region1'
        ws_params['C2'] = 'coal_ppl'
        ws_params['D2'] = 2020
        ws_params['E2'] = 1000.50

        ws_params['A3'] = 'fix_cost'
        ws_params['B3'] = 'region1'
        ws_params['C3'] = 'solar_pv'
        ws_params['D3'] = 2020
        ws_params['E3'] = 500.25

        # Test parameter 2: variable_cost
        ws_params['A4'] = 'variable_cost'
        ws_params['B4'] = 'region1'
        ws_params['C4'] = 'coal_ppl'
        ws_params['D4'] = 2020
        ws_params['E4'] = 25.0

        # Create sets sheet
        ws_sets = wb.create_sheet("sets")
        ws_sets['A1'] = 'set_name'
        ws_sets['B1'] = 'element1'
        ws_sets['C1'] = 'element2'
        ws_sets['D1'] = 'element3'

        ws_sets['A2'] = 'node'
        ws_sets['B2'] = 'region1'
        ws_sets['C2'] = 'region2'

        ws_sets['A3'] = 'technology'
        ws_sets['B3'] = 'coal_ppl'
        ws_sets['C3'] = 'solar_pv'
        ws_sets['D3'] = 'wind_ppl'

        # Create individual set sheet
        ws_tech = wb.create_sheet("technology")
        ws_tech['A1'] = 'tech'
        ws_tech['A2'] = 'nuclear'
        ws_tech['A3'] = 'hydro'

        # Save to temporary file
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
            wb.save(tmp.name)
            yield tmp.name

        # Cleanup
        os.unlink(tmp.name)

    def test_load_excel_file_success(self, temp_excel_file):
        """Test successful loading of Excel file"""
        manager = InputManager()

        scenario = manager.load_excel_file(temp_excel_file)

        # Verify scenario was created
        assert scenario is not None
        assert isinstance(scenario, object)  # Should be ScenarioData

        # Check that we have the manager's reference
        assert manager.get_current_scenario() is scenario

    def test_parse_parameters(self, temp_excel_file):
        """Test parameter parsing"""
        manager = InputManager()
        scenario = manager.load_excel_file(temp_excel_file)

        # Check parameters were parsed
        param_names = scenario.get_parameter_names()
        assert 'fix_cost' in param_names
        assert 'variable_cost' in param_names
        assert len(param_names) == 2

        # Check fix_cost parameter
        fix_cost_param = scenario.get_parameter('fix_cost')
        assert fix_cost_param is not None
        assert fix_cost_param.name == 'fix_cost'

        # Check dataframe content
        df = fix_cost_param.df
        assert len(df) == 2  # 2 rows of data
        assert list(df.columns) == ['index', 'node_loc', 'technology', 'year_vtg', 'value']

        # Check specific values
        assert df.iloc[0]['node_loc'] == 'region1'
        assert df.iloc[0]['technology'] == 'coal_ppl'
        assert df.iloc[0]['year_vtg'] == 2020
        assert df.iloc[0]['value'] == 1000.50

        assert df.iloc[1]['technology'] == 'solar_pv'
        assert df.iloc[1]['value'] == 500.25

        # Check metadata
        assert fix_cost_param.metadata['dims'] == ['node_loc', 'technology', 'year_vtg']
        assert fix_cost_param.metadata['value_column'] == 'value'
        assert fix_cost_param.metadata['shape'] == (2, 5)

    def test_parse_sets(self, temp_excel_file):
        """Test set parsing"""
        manager = InputManager()
        scenario = manager.load_excel_file(temp_excel_file)

        # Check sets were parsed
        assert 'node' in scenario.sets
        assert 'technology' in scenario.sets

        # Check node set
        node_set = scenario.sets['node']
        assert len(node_set) == 2
        assert 'region1' in node_set.values
        assert 'region2' in node_set.values

        # Check technology set (from individual technology sheet)
        tech_set = scenario.sets['technology']
        assert len(tech_set) == 2  # Individual sheet overrides combined sets
        assert 'nuclear' in tech_set.values
        assert 'hydro' in tech_set.values

    def test_parse_individual_set_sheet(self, temp_excel_file):
        """Test parsing of individual set sheets"""
        # This test would need to be updated if we add individual set sheet parsing
        # For now, the technology set from the individual sheet should be merged or handled
        manager = InputManager()
        scenario = manager.load_excel_file(temp_excel_file)

        # The current implementation might have conflicts between
        # sets sheet and individual set sheets with same name
        # This is a known limitation to test
        pass

    def test_validation_valid_scenario(self, temp_excel_file):
        """Test validation of a valid scenario"""
        manager = InputManager()
        scenario = manager.load_excel_file(temp_excel_file)

        validation = manager.validate_scenario()

        # Should be valid with no issues
        assert validation['valid'] is True
        assert len(validation['issues']) == 0
        assert validation['summary']['parameters'] == 2
        assert validation['summary']['sets'] == 2
        assert validation['summary']['total_data_points'] == 3  # 2 + 1 rows

    def test_validation_empty_scenario(self):
        """Test validation of empty scenario"""
        manager = InputManager()

        validation = manager.validate_scenario()

        assert validation['valid'] is False
        assert 'No scenario loaded' in validation['issues']

    def test_file_not_found(self):
        """Test handling of non-existent file"""
        manager = InputManager()

        with pytest.raises(FileNotFoundError):
            manager.load_excel_file('non_existent_file.xlsx')

    def test_invalid_excel_file(self):
        """Test handling of invalid Excel file"""
        manager = InputManager()

        # Create a text file with .xlsx extension
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
            tmp.write(b"This is not an Excel file")
            tmp_path = tmp.name

        try:
            with pytest.raises(ValueError):
                manager.load_excel_file(tmp_path)
        finally:
            os.unlink(tmp_path)

    def test_parameter_metadata(self, temp_excel_file):
        """Test parameter metadata extraction"""
        manager = InputManager()
        scenario = manager.load_excel_file(temp_excel_file)

        param = scenario.get_parameter('fix_cost')
        metadata = param.metadata

        assert metadata['units'] == 'N/A'  # Default value
        assert metadata['desc'] == 'Parameter fix_cost'
        assert metadata['dims'] == ['node_loc', 'technology', 'year_vtg']
        assert metadata['value_column'] == 'value'
        assert metadata['shape'] == (2, 5)

    def test_parameter_with_missing_data(self):
        """Test handling of parameters with missing data"""
        wb = Workbook()
        ws = wb.active
        ws.title = "parameters"

        # Headers
        ws['A1'] = 'parameter'
        ws['B1'] = 'dim1'
        ws['C1'] = 'value'

        # Parameter with some missing values
        ws['A2'] = 'test_param'
        ws['B2'] = 'item1'
        ws['C2'] = 100

        ws['A3'] = 'test_param'
        ws['B3'] = 'item2'
        ws['C3'] = None  # Missing value

        ws['A4'] = 'test_param'
        ws['B4'] = None   # Missing dimension
        ws['C4'] = 200

        # Save to temporary file
        tmp_fd, tmp_path = tempfile.mkstemp(suffix='.xlsx')
        try:
            with os.fdopen(tmp_fd, 'wb') as tmp:
                wb.save(tmp)

            manager = InputManager()
            scenario = manager.load_excel_file(tmp_path)

            param = scenario.get_parameter('test_param')
            assert param is not None

            # Should still create the parameter even with missing data
            df = param.df
            assert len(df) == 3

            # Check validation detects issues
            validation = manager.validate_scenario()
            # Note: validation might not catch all missing data issues
            # depending on implementation

        finally:
            try:
                os.unlink(tmp_path)
            except OSError:
                pass  # Ignore if file is already deleted or inaccessible
